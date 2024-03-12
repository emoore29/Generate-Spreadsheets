from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
from tabulate import tabulate
import time
import json
from datetime import date

# Define global variables
viq_file_path = None
# Get the current date
current_date = datetime.now().date()
# find the date of next Monday
days_until_monday = (7 - current_date.weekday()) % 7
next_monday = current_date + timedelta(days=days_until_monday)

weekly_availability = {
    'Monday': {'date': next_monday, 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '0.5',
        'five_day': '',
    }, 'mdpi': '<=10k'},
    'Tuesday': {'date': next_monday + timedelta(days=1), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '0.5',
        'five_day': '',
    }, 'mdpi': '<=15k'},
    'Wednesday': {'date': next_monday + timedelta(days=2), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '0.5',
        'five_day': '',
    }, 'mdpi': '<=20k'},
    'Thursday': {'date': next_monday + timedelta(days=3), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '',
        'five_day': '',
    }, 'mdpi': '<=10k'},
    'Friday': {'date': next_monday + timedelta(days=4), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '',
        'five_day': '',
    }, 'mdpi': '<=5k'},
    'Saturday': {'date': next_monday + timedelta(days=5), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '',
        'five_day': '',
    }, 'mdpi': ''},
    'Sunday': {'date': next_monday + timedelta(days=6), 'viq': {
        'live_am': '',
        'live_pm': '',
        'overnight': '',
        'two_day': '',
        'five_day': '',
    }, 'mdpi': ''},
}

# Convert dictionary to list of lists for tabulate
def convert_to_table():
    table_data = []
    for day, details in weekly_availability.items():
        # Extract viq details
        viq_details = details['viq']
        # Construct VIQ string based on non-empty details
        viq_string = ", ".join(f"{key.replace('_', ' ').title()}: {value}" for key, value in viq_details.items() if value)
        table_data.append([day, details['date'].strftime('%d %B %Y'), viq_string, details['mdpi']])
        
    return table_data

def update_viq():
    global viq_file_path
    
    print("Opening next week's availability file...")
    # Add a 1-second delay so user can read the above.
    time.sleep(1)
    os.startfile(viq_file_path)
    
    updated = input("Have you updated the file? y/n: ") 
    
    if updated:
        # Load the updated Excel file
        wb_updated = load_workbook(viq_file_path)
        sheet_updated = wb_updated.active
        
        # Define a dictionary to map each day of the week to its corresponding column in the Excel sheet
        day_columns = {'Monday': 'B', 'Tuesday': 'D', 'Wednesday': 'F', 'Thursday': 'H', 'Friday': 'J'}
        
        # Iterate over the days of the week and update the weekly_availability dictionary
        for day, details in weekly_availability.items():
            # Check if the current day is Saturday
            if day == 'Saturday':
                break
            
            # Get the corresponding column letter from the day_to_column dictionary
            cell_column = day_columns.get(day)
            
            # Assign values to variables
            live_am = sheet_updated[f"{cell_column}6"].value
            live_pm = sheet_updated[f"{cell_column}7"].value
            overnight = sheet_updated[f"{cell_column}8"].value
            two_day = sheet_updated[f"{cell_column}9"].value
            five_day = sheet_updated[f"{cell_column}10"].value
            
            # Assign variable values to weekly_availability details
            details['viq']['live_am'] = live_am
            details['viq']['live_pm'] = live_pm
            details['viq']['overnight'] = overnight
            details['viq']['two_day'] = two_day
            details['viq']['five_day'] = five_day
        
        print("Next week's VIQ availability updated successfully: ")
        print(tabulate(convert_to_table(), headers=["Day", "Date", "VIQ", "MDPI"], tablefmt="grid"))
    else:
        print("Please update the file and return to this program.")
        
           

def create_new_spreadsheet():
    global viq_file_path
    # Create new VIQ spreadsheet for next week and delete old one.
    template_path = "C:/Users/emoor/OneDrive/Documents/VIQ/Availability/Emmaline Moore - Availability Template.xlsx"
    output_folder = "C:/Users/emoor/OneDrive/Documents/VIQ/Availability"

    # load availability template
    wb = load_workbook(template_path)
        
    # open the currently active sheet in the workbook
    sheet = wb.worksheets[0]

    # Update date in cell B4 to next monday's date
    sheet["B4"].value=next_monday

    # Save the updated template to a new Excel file
    new_file_name = f"Emmaline Moore - Availability Week Commencing {next_monday.strftime('%d %B %Y')}.xlsx"
    viq_file_path = os.path.join(output_folder, new_file_name)
    wb.save(viq_file_path)
    


create_new_spreadsheet()
# Print welcome message
print(f"Hello! Welcome to Auto Availability! It's currently {current_date}. Next Monday is {next_monday}. A spreadsheet for your VIQ availability next week has been created at {viq_file_path}. Here is your default schedule: ")
print(tabulate(convert_to_table(), headers=["Day", "Date", "VIQ", "MDPI"], tablefmt="grid"))

update_viq_input = input("Do you want to update your VIQ availability? y/n: ")

if update_viq_input == "y":
    update_viq()

updated_mdpi = input("Have you updated MDPI on SuSy? y/n: ")

if updated_mdpi == "y":
    for day, details in weekly_availability.items():
        details['mdpi'] = ''
    
    mdpi_input = input("Next week has been cleared. Please list any days that you have added a wordcount for (e.g. mon-15). Remember to account for your availability time and when work is likely to be allocated/due: ")
    
    # Define a mapping between abbreviated and full day names
    day_mapping = {'sun': 'Sunday', 'mon': 'Monday', 'tue': 'Tuesday', 'wed': 'Wednesday', 'thu': 'Thursday', 'fri': 'Friday', 'sat': 'Saturday'}
    
    # Split the input by commas to get individual day-word count pairs
    day_word_count_pairs = mdpi_input.split(',')
    
    for pair in day_word_count_pairs:
        # Split each pair into day and word count
        day_abbr, word_count = pair.split('-')
        
        # Convert abbreviated day name to full day name
        day_full = day_mapping.get(day_abbr)
        
        # Update the MDPI value for the corresponding day in the weekly_availability dictionary
        weekly_availability[day_full]['mdpi'] = f"<={word_count}k"
    
    print("Next week's MDPI availability updated successfully: ")
    print(tabulate(convert_to_table(), headers=["Day", "Date", "VIQ", "MDPI"], tablefmt="grid"))
    
print("This is your schedule for next week. If you want to make edits, please run the program again because I haven't written any ability to edit from this point onward.")
print(tabulate(convert_to_table(), headers=["Day", "Date", "VIQ", "MDPI"], tablefmt="grid"))

print("Saving weekly availability as a JSON file...")

# Custom JSON encoder that supports datetime.date objects
class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, date):
            return obj.isoformat()  # Serialize date objects as ISO formatted strings
        return super().default(obj)

# Path to Quick Lists json location
json_file_path = "C:/Users/emoor/OneDrive/Documents/GitHub/Quick-Lists/json/weekly_availability.json"

# Save weekly_availability as a JSON file
with open(json_file_path, 'w') as json_file: # Opens weekly_availability.json in write mode. Creates if it doesn't exist. 'with' will close file after json.dump is executed.
    json.dump(weekly_availability, json_file, cls=DateEncoder) # Turns weekly_availability into json format. Writes to json_file.

print("Saved as JSON file to Quick Lists. You will need to push this change to GitHub if you want Quick Lists to update.")
